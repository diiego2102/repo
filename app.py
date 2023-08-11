import streamlit as st
import os
from PIL import ImageOps
from PIL import Image as PILImage
#from office365.runtime.auth.authentication_context import AuthenticationContext
#from office365.runtime.auth.client_credential import ClientCredential
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
import math
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break
import glob 
import os
from datetime import datetime
from datetime import date
import io
import pandas as pd
import locale
locale.setlocale(locale.LC_TIME, 'en_US.UTF-8')
from funciones_sharepoint import * 
# Cargar imagen
st.set_page_config(page_title="Control Pérdidas - Valorizaciones APP", page_icon="calavera.jfif")
st.title("Control Pérdidas - Valorizaciones APP")
st.caption("Control de pérdidas")
suministro_selection = st.text_input("Ingrese Número de Suministro")
choice = ["1 Carga de Acta y Fotografías"]
choice_value = st.sidebar.selectbox("Seleccionar el proceso", choice)
rot_image = None
# Fechas de trabajo
fecha_interv = st.date_input("Ingresa fecha de intervención")
fecha_interv = fecha_interv.strftime('%Y-%m-%d')
mes_aux = date.fromisoformat(fecha_interv).strftime('%B').replace(".","").capitalize()
fecha_ruta = f'{date.fromisoformat(fecha_interv).year}/{int(date.fromisoformat(fecha_interv).month)} {mes_aux}/{int(date.fromisoformat(fecha_interv).day)}'
#RUTA DE SHAREPOINT
folder_valorizacion = f'{folder_url}/Valorizacion Excel/{fecha_ruta}'
folder_fotos = f'{folder_url}/Actas Fotos/{fecha_ruta}'

# Funcion rotacion
def btn_rotate_image(img,degrees):
    rotated_image = img.rotate(degrees, expand=True)
    return rotated_image

if choice_value == "1 Carga de Acta y Fotografías":
    st.title(choice[0])

if choice_value == "1 Carga de Acta y Fotografías":
    carga_options = ['Cargar Acta de intervención','Cargar Actas Fotográficas','Acta Fotográfica XLSX']
    carga_box = st.radio("Tipo de Carga",carga_options)

    if carga_box == carga_options[0]:
        if "rotated_image" not in st.session_state:
            st.session_state["rotated_image"] = None
        uploaded_file = st.file_uploader("Cargar imagen de Acta de Intervención", type=["jpg", "jpeg", "png"])
        
        # Mostrar preview de la imagen y permitir al usuario ajustar la rotación
        if uploaded_file is not None:
            # Abrir imagen con Pillow
            image = PILImage.open(uploaded_file)
            st.sidebar.subheader("Rota las imágenes")
            # Rotar imagen según la orientación EXIF
            image = ImageOps.exif_transpose(image) 
            st.image(image, caption="Imagen original", use_column_width=True)

            if "degrees" not in st.session_state:
                st.session_state["degrees"] = 0

            col1, col2= st.sidebar.columns(2)
            if col1.button('↪️'):
                # rot_image = image.rotate(rotation)
                st.session_state["degrees"] += 90
                rotated_image = rotation_function().btn_rotate_image(image, st.session_state["degrees"])
                st.session_state["rotated_image"] = rotated_image
                st.image(rotated_image, caption="Imagen girada")
                
            if col2.button('↩️'):
                st.session_state["degrees"] -= 90
                rotated_image = rotation_function().btn_rotate_image(image, st.session_state["degrees"])
                st.session_state["rotated_image"] = rotated_image
                st.image(rotated_image, caption="Imagen girada")

            if st.sidebar.button("Guardar Acta de Intervención"):
                if st.session_state["rotated_image"] is not None:
                    image_ref = f'Acta de intervención sum. {suministro_selection}'
                    filename = f"{image_ref}{'.jpeg'}"
                    sharepoint(sharepoint_url_site,folder_fotos,sharepoint_email,sharepoint_password).upload_image(filename,st.session_state["rotated_image"])                 
            else:
                pass
    elif carga_box == carga_options[1]:
        image_labels = {}
        label_options = ['1_MEDIDOR INTERVENIDO','2_AGREGAR CAUSALIDAD','3_SERVICIO CORTADO','4_FASE R','5_FASE S','6_FASE T','7_VIVIENDA','8_OTROS']
        uploaded_images = st.file_uploader("Cargar las imagenes de la Intervención", type=["jpg", "jpeg", "png"], accept_multiple_files=True)
        if uploaded_images is not None:
            num_columns = 3
            num_images = len(uploaded_images)
            num_rows = math.ceil(num_images / num_columns)
   

            rotation_options = [-180, -90, 0, 90, 180]
            def rotate_image(image, rotation):
                return image.rotate(rotation, expand=True)
            
            for i in (range(num_rows)):
                cols = st.columns(num_columns)
                for j in range(num_columns):
                    index = i * num_columns + j
                    if index < num_images:
                        image_file = uploaded_images[index]
                        image_data = image_file.read()
                        cols[j].image(image_data, use_column_width=True, caption=f"Imagen_N°_{index}")

            selected_index = st.sidebar.selectbox("Selecciona la imagen para renombrar y rotar:", range(num_images), index=0)
            selected_image = uploaded_images[selected_index]
            rotation = st.sidebar.selectbox("Rota la imagen:", rotation_options, index=0)
            # Add labels for the uploaded images
            label_key = f"label_{index}"
            label_value = st.sidebar.selectbox("Selecciona label:", label_options, key=label_key)
            image_labels[index] = label_value

            # Add labels for the uploaded images
            #label_key = f"label_{index}"
            #label_value = st.sidebar.selectbox("Selecciona label:", label_options, key=label_key)
            #if index == 1:
                #label_value = st.sidebar.text_input('Ingresa la causalidad:')
                #label_value = fr'2_{label_value}'
            #else:
                #image_labels[index] = label_value

            st.sidebar.subheader("Antes")
            selected_image_file = PILImage.open(selected_image)

            if st.sidebar.button("Rotar y guardar"):
                st.sidebar.image(selected_image, use_column_width=True, caption=selected_image)
                rotated_image = rotate_image(selected_image_file, rotation)
                st.sidebar.subheader("Después")
                st.sidebar.image(rotated_image, use_column_width=True, caption=selected_image)
                image_extension = os.path.splitext(selected_image.name)[1]
                # image_path = os.path.join("uploads", f"{label_value}{image_extension}")
                #image_path = os.path.join(f'D:\Actas_Fotograficas\{suministro_selection}_output', f"{label_value}{image_extension}")
                #os.makedirs(f'D:\Actas_Fotograficas\{suministro_selection}_output',exist_ok=True)
                #img_bytes = rotated_image.tobytes()
                img_bytes = io.BytesIO()
                img_bytes.seek(0)
                st.download_button(
                    label=f" Descarga {label_value}_{suministro_selection}",
                    data=img_bytes,
                    file_name=f"{label_value}_{suministro_selection}{image_extension}", 
                    mime='image/jpeg'
                )
                # rotated_image.save(img_bytes, format='JPEG')
                # img_bytes = img_bytes.getvalue()
                # st.sidebar.download_button(label='Descargar Imagen', data=img_bytes, file_name=f"{label_value}_S{suministro_selection}{image_extension}", mime='image/jpeg')
                #rotated_image.save(image_path)
                st.sidebar.success(f"Imagen rotada guardada como {label_value}{image_extension}")


        # Display image labels
        st.sidebar.subheader("Labels de la Imagen")
        for index, label in image_labels.items():
            st.sidebar.write(f"Imagen_N°_{index}: {label}")



    elif carga_box == carga_options[2]:
        ctx = sharepoint(sharepoint_url_site,folder_fotos,sharepoint_email,sharepoint_password)._auth()
        target_folder = ctx.web.get_folder_by_server_relative_url(folder_fotos)
        # Create a new workbook
        wb = openpyxl.Workbook()
        buffer = BytesIO()
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_view.showGridLines = False

        uploaded_images = st.file_uploader("Cargar las imagenes de la Intervención", type=["jpg", "jpeg", "png"], accept_multiple_files=True)

        if uploaded_images:
            # Ordenar las imágenes por nombre de archivo
            uploaded_images = sorted(uploaded_images, key=lambda image: image.name)
        st.write(uploaded_images)
        if uploaded_images is not None:
            num_columns = 3
            num_images = len(uploaded_images)
            num_rows = math.ceil(num_images / num_columns)

        selected_index = st.sidebar.selectbox("Selecciona la cantidad de imágenes para :", range(1,21), index=0)
        image_paths = []
        for uploaded_file in uploaded_images:
                image_paths.append(uploaded_file.name)
        image_paths.sort()
        st.write(image_paths)
        if selected_index <=6:
            cells = ['A1','E1','I1','A17','E17','I17']
            cells_name = ['A15','E15','I15','A32','E32','I32']

            # Imagenes
            ws.merge_cells(start_row=1, start_column=1, end_row=14, end_column=3) #A1
            ws.merge_cells(start_row=1, start_column=5, end_row=14, end_column=7) #E1
            ws.merge_cells(start_row=1, start_column=9, end_row=14, end_column=11) #I1

            ws.merge_cells(start_row=17, start_column=1, end_row=31, end_column=3) #A17
            ws.merge_cells(start_row=17, start_column=5, end_row=31, end_column=7) #E17
            ws.merge_cells(start_row=17, start_column=9, end_row=31, end_column=11) #I17

            #Comentario
            ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=3) #A16
            ws.merge_cells(start_row=15, start_column=5, end_row=15, end_column=7) #E16
            ws.merge_cells(start_row=15, start_column=9, end_row=15, end_column=11) #I16

            ws.merge_cells(start_row=32, start_column=1, end_row=32, end_column=3) #A32
            ws.merge_cells(start_row=32, start_column=5, end_row=32, end_column=7) #E32
            ws.merge_cells(start_row=32, start_column=9, end_row=32, end_column=11) #I32
            # print(uploaded_images)
            for img_, cell in zip(uploaded_images,cells):
                img = Image(img_)
                img.width = 190
                img.height = 281
                ws.add_image(img,cell)
            
            for img_path, cell_name in zip(image_paths,cells_name):
                name = os.path.basename(img_path).split('_')[1] #.split('.')[0]
                ws[cell_name] = name
                ws[cell_name].alignment = Alignment(horizontal='center', vertical='center')
        
        elif selected_index >= 7 and selected_index <=12:
            cells = ['A1','E1','I1','A17','E17','I17','A33','E33','I33','A49','E49','I49']
            cells_name = ['A16','E16','I16','A32','E32','I32','A47','E47','I47','A63','E63','I63']

            # Imagenes
            ws.merge_cells(start_row=1, start_column=1, end_row=14, end_column=3) #A1
            ws.merge_cells(start_row=1, start_column=5, end_row=14, end_column=7) #E1
            ws.merge_cells(start_row=1, start_column=9, end_row=14, end_column=11) #I1

            ws.merge_cells(start_row=17, start_column=1, end_row=31, end_column=3) #A17
            ws.merge_cells(start_row=17, start_column=5, end_row=31, end_column=7) #E17
            ws.merge_cells(start_row=17, start_column=9, end_row=31, end_column=11) #I17

            ws.merge_cells(start_row=33, start_column=1, end_row=46, end_column=3) #A33
            ws.merge_cells(start_row=33, start_column=5, end_row=46, end_column=7) #E33
            ws.merge_cells(start_row=33, start_column=9, end_row=46, end_column=11) #I33

            ws.merge_cells(start_row=49, start_column=1, end_row=62, end_column=3) #A49
            ws.merge_cells(start_row=49, start_column=5, end_row=62, end_column=7) #E49
            ws.merge_cells(start_row=49, start_column=9, end_row=62, end_column=11) #I49

            #Comentario
            ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=3) #A16
            ws.merge_cells(start_row=15, start_column=5, end_row=15, end_column=7) #E16
            ws.merge_cells(start_row=15, start_column=9, end_row=15, end_column=11) #I16

            ws.merge_cells(start_row=32, start_column=1, end_row=32, end_column=3) #A32
            ws.merge_cells(start_row=32, start_column=5, end_row=32, end_column=7) #E32
            ws.merge_cells(start_row=32, start_column=9, end_row=32, end_column=11) #I32

            ws.merge_cells(start_row=47, start_column=1, end_row=47, end_column=3) #A47
            ws.merge_cells(start_row=47, start_column=5, end_row=47, end_column=7) #E47
            ws.merge_cells(start_row=47, start_column=9, end_row=47, end_column=11) #I47

            ws.merge_cells(start_row=63, start_column=1, end_row=63, end_column=3) #A63
            ws.merge_cells(start_row=63, start_column=5, end_row=63, end_column=7) #E63
            ws.merge_cells(start_row=63, start_column=9, end_row=63, end_column=11) #I63
            # print(uploaded_images)

            for img_, cell in zip(uploaded_images,cells):
                img = Image(img_)
                img.width = 190
                img.height = 281
                ws.add_image(img,cell)
            
            for img_path, cell_name in zip(image_paths,cells_name):
                name = os.path.basename(img_path).split('_')[1] #.split('.')[0]
                ws[cell_name] = name
                ws[cell_name].alignment = Alignment(horizontal='center', vertical='center')
        else:
            pass
        ####timestamp
        today = datetime.now()
        d1 = today.strftime("%d%m%y_%H%M")
        # file = io.BytesIO()
        wb.save(buffer)
        # file.seek(0)
        # st.download_button(
        #     label='Descarga Acta Fotográfica',
        #     data=file,
        #     file_name=f'Acta Fotografica {suministro_selection}.xlsx',
        #     mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
        file_content = buffer.getvalue()
        response = target_folder.upload_file(f'Acta Fotografica {suministro_selection}.xlsx', file_content).execute_query()
    else:
        pass
else:
    pass 
