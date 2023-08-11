#from office365.runtime.auth.authentication_context import AuthenticationContext
#from office365.runtime.auth.client_credential import ClientCredential
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
import os
import tempfile
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import requests
from urllib3.exceptions import InsecureRequestWarning
from io import BytesIO
from openpyxl import load_workbook
import streamlit as st
import streamlit as ReportThread
from streamlit_server_state import server_state
import os
from PIL import ImageOps
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
from concurrent.futures import ThreadPoolExecutor, as_completed
 
# Suppress only the single warning from urllib3 needed.
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

sharepoint_email = 'dgallo@luzdelsur.com.pe'
sharepoint_password = 'Mason2100$'
sharepoint_url_site = 'https://luzdelsurlds.sharepoint.com/sites/CoordinacinBalanceydeteccindefraude'
sharepoint_doc_library = 'Documentos%20compartidos'
sharepoint_site_name = 'CoordinacinBalanceydeteccindefraude'
folder_url = f'/sites/{sharepoint_site_name}/{sharepoint_doc_library}/General/11 Valorizaciones CNR'
# folder_valorizacion = f'{folder_url}/Valorizacion Excel'
# folder_fotos = f'{folder_url}/Actas Fotos'

class sharepoint():
    def __init__(self,url_site,url_folder,user,password):
        self.url_site = url_site
        self.url_folder = url_folder
        self.user = user
        self.password = password

    def _auth(self):
        credential = UserCredential(self.user,self.password)
        ctx = ClientContext(self.url_site).with_credentials(credential)
        ctx.ssl_cert_validation = ssl.CERT_NONE
        return ctx
    
    def list_folder_url(self):
        ctx = self._auth()
        folder = ctx.web.get_folder_by_server_relative_url(self.url_folder)#.expand("Files")
        files = folder.files.get().execute_query()
        folders = folder.folders.get().execute_query()
        lista_folders = []
        for folder in folders:
            # print("Carpeta: " + folder.properties["ServerRelativeUrl"])<<
            lista_folders.append(folder.properties["ServerRelativeUrl"])
        lista_files = []
        for i_file in files:
            # print("Archivo: " + file.properties["ServerRelativeUrl"])
            lista_files.append(i_file.properties["ServerRelativeUrl"])
        return lista_files,lista_folders      
 
    def download_files_v1(self,file_name,outpath):
        url = f"{self.url_site}/_api/web/GetFileByServerRelativeUrl({self.url_folder} + f'/{file_name}')/$value"
        auth = requests.auth.HTTPBasicAuth(self.user, self.password)
        # Descargar el archivo
        response = requests.get(url, auth=auth,verify=False)
        with open(fr'{outpath}\hola.xlsx', "wb") as f:
            f.write(response.content)       

    def download_files_v2(self,file_name,output):
        ctx = self._auth()
        url_file = f'{self.url_folder}/{file_name}'
        file = File.open_binary(ctx, url_file)
        file_bytes = BytesIO(file.content)
        workbook = load_workbook(file_bytes)
        sheet_names = workbook.sheetnames
        workbook.save(fr'{output}/{file_name}')
        print(f'Se guardó el archivo en {output}')
    
    def upload_to_sharepoint(self,upload_file_lst: list):
        ctx = self._auth()
        target_folder_url = f'{self.url_folder}'
        target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)

        with ThreadPoolExecutor(max_workers=5) as executor:
            for file_name in upload_file_lst:
                with open(file_name, 'rb') as content_file:
                    file_content = content_file.read()
                    futures = executor.submit(target_folder.upload_file, f'{target_folder_url}/{os.path.basename(file_name)}', file_content)
                    if as_completed(futures):
                        futures.result().execute_query()
        
        print(f'Se subio el archivo {os.path.basename(file_name)}')

    def upload_image(self,image_filename,image):
        # Guardar imagen modificada en un archivo temporal
        temp_image_path = image_filename
        image.save(temp_image_path)

        ctx = self._auth()
        target_folder_url = f'{self.url_folder}'
        target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)

        with open(temp_image_path, 'rb') as f:
            content = f.read()
            response = target_folder.upload_file(os.path.basename(temp_image_path), content).execute_query()

        # Eliminar archivo temporal
        os.remove(temp_image_path)
        print(f'El archivo se subió en {target_folder_url}')
    
class rotation_function():
    def btn_rotate_image(self,img,degrees):
        rotated_image = img.rotate(degrees, expand=True)
        return rotated_image
    
    def get_session_state(self,):
        session_id = ReportThread.get_report_ctx().session_id
        session_info = server_state.get_current()._get_session_info(session_id)
        if session_info is None:
            session_state = {"degrees": 0}
        else:
            session_state = session_info.session.session_state
        return session_state

    def set_session_state(self,session_state):
        session_id = ReportThread.get_report_ctx().session_id
        session_info = server_state.get_current()._get_session_info(session_id)
        if session_info is not None:
            session_info.session.session_state = session_state
